from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS_FULL: List[str] = [
    "Soru Metni",
    "Seçenek Sayısı",
    "Seçenek 1", "Seçenek 2", "Seçenek 3", "Seçenek 4", "Seçenek 5",
    "Doğru Seçenek",
    "Zorluk Derecesi",
    "Departman",
    "Eğitim",
    "Konu",
    "Amaç",
    "Hazırlayan",
]

HEADERS_MINIMAL: List[str] = [
    "Soru Metni",
    "Seçenek Sayısı",
    "Seçenek 1", "Seçenek 2", "Seçenek 3", "Seçenek 4", "Seçenek 5",
    "Doğru Seçenek",
    "Zorluk Derecesi",
]

_LETTER_TO_NUM = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
_DIF_MAP = {"Kolay": 1, "Orta": 3, "Zor": 5}
_DEFAULT_PREPARER = "egitim.yonetici"


@dataclass
class ExcelMeta:
    zorluk_derecesi: str = "Orta"
    departman: str = ""
    egitim: str = ""
    konu: str = ""
    amac: str = ""
    hazirlayan: str = _DEFAULT_PREPARER


def _options_list(item: Dict[str, Any]) -> List[str]:
    opts = item.get("options")
    if not opts:
        return []

    if isinstance(opts, dict):
        def rank(k: Any) -> int:
            s = str(k).strip().upper()
            if s in "ABCDE":
                return "ABCDE".index(s)
            if s.isdigit():
                return int(s) - 1
            return 999

        keys_sorted = sorted(opts.keys(), key=rank)
        return [str(opts[k]).strip() for k in keys_sorted][:5]

    if isinstance(opts, list):
        return [str(x).strip() for x in opts][:5]

    return []


def _correct_number(item: Dict[str, Any]) -> Optional[int]:
    qtype = str(item.get("type", "")).lower().strip()
    corr = item.get("answer") if item.get("answer") is not None else item.get("correct")

    if qtype in ("true_false", "tf"):
        if isinstance(corr, str):
            s = corr.strip().lower()
            if s in ("doğru", "dogru"):
                return 1
            if s in ("yanlış", "yanlis"):
                return 2
        if isinstance(corr, int) and corr in (1, 2):
            return corr
        return None

    if qtype in ("fill", "blank", "fill_blank"):
        return 1 if item.get("answer") else None

    if isinstance(corr, str):
        if corr.upper() in _LETTER_TO_NUM:
            return _LETTER_TO_NUM[corr.upper()]
        if corr.strip().isdigit():
            return int(corr.strip())
    if isinstance(corr, int):
        return corr
    return None


def _difficulty(item: Dict[str, Any], meta: ExcelMeta) -> int:
    d = item.get("difficulty") or item.get("zorluk") or meta.zorluk_derecesi
    if isinstance(d, int):
        return d
    s = str(d).strip()
    if s.isdigit():
        return int(s)
    return _DIF_MAP.get(s.capitalize(), 3)


def _base_row(item: Dict[str, Any], meta: ExcelMeta) -> List[Any]:
    qtype = str(item.get("type", "")).lower().strip()
    soru = str(item.get("question", "")).strip()
    options = _options_list(item)

    if qtype in ("mcq", "multiple_choice"):
        secenek_sayisi = len(options)
        opt1, opt2, opt3, opt4, opt5 = (options + [""] * 5)[:5]
    elif qtype in ("true_false", "tf"):
        secenek_sayisi = 2
        opt1, opt2, opt3, opt4, opt5 = "Doğru", "Yanlış", "", "", ""
    elif qtype in ("fill", "blank", "fill_blank"):
        ans = str(item.get("answer", "")).strip()
        secenek_sayisi = 1 if ans else 0
        opt1, opt2, opt3, opt4, opt5 = ans, "", "", "", ""
    else:
        secenek_sayisi = 0
        opt1 = opt2 = opt3 = opt4 = opt5 = ""

    return [
        soru,
        secenek_sayisi,
        opt1, opt2, opt3, opt4, opt5,
        _correct_number(item),
        _difficulty(item, meta),
    ]


def _error_row(item: Dict[str, Any], meta: ExcelMeta, include_metadata: bool) -> List[Any]:
    err = str(item.get("error", "")).strip()
    raw_prev = str(item.get("raw_preview", "")).strip()
    qtype = str(item.get("type", "")).lower().strip()
    msg = f"[ERROR] {qtype}: {err}"
    if raw_prev:
        msg += f" | raw: {raw_prev[:250]}"

    row: List[Any] = [msg, 0, "", "", "", "", "", "", _difficulty(item, meta)]
    if include_metadata:
        row += [meta.departman, meta.egitim, meta.konu, meta.amac, meta.hazirlayan]
    return row


def quiz_to_rows(
    quiz: List[Dict[str, Any]],
    meta: ExcelMeta,
    include_metadata: bool = True,
) -> List[List[Any]]:
    rows: List[List[Any]] = []

    for item in quiz:
        if isinstance(item, dict) and item.get("error"):
            rows.append(_error_row(item, meta, include_metadata))
            continue

        row = _base_row(item, meta)
        if include_metadata:
            row += [meta.departman, meta.egitim, meta.konu, meta.amac, meta.hazirlayan]
        rows.append(row)

    return rows


_HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_COL_WIDTHS_FULL = {
    "A": 60, "B": 16,
    "C": 26, "D": 26, "E": 26, "F": 26, "G": 26,
    "H": 22, "I": 18,
    "J": 20, "K": 22, "L": 20, "M": 20, "N": 20,
}
_COL_WIDTHS_MINIMAL = {
    "A": 60, "B": 16,
    "C": 26, "D": 26, "E": 26, "F": 26, "G": 26,
    "H": 22, "I": 18,
}
_WRAP_COLS_FULL = {"A", "C", "D", "E", "F", "G", "M"}
_WRAP_COLS_MINIMAL = {"A", "C", "D", "E", "F", "G"}


def _write_header(ws, headers: List[str]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _write_metrics_sheet(wb, metrics: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Metrics")
    _write_header(ws, ["Metric", "Value"])

    for k in sorted(metrics.keys()):
        v = metrics.get(k)
        if isinstance(v, (dict, list, tuple)):
            val = str(v)
        else:
            val = "" if v is None else str(v)
        ws.append([str(k), val])

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def export_quiz_to_xlsx(
    quiz: List[Dict[str, Any]],
    meta: Optional[ExcelMeta] = None,
    out_path: str = "quiz.xlsx",
    sheet_name: str = "Quiz",
    metrics: Optional[Dict[str, Any]] = None,
    include_metadata: bool = True,
) -> str:
    meta = meta or ExcelMeta()
    headers = HEADERS_FULL if include_metadata else HEADERS_MINIMAL
    col_widths = _COL_WIDTHS_FULL if include_metadata else _COL_WIDTHS_MINIMAL
    wrap_cols = _WRAP_COLS_FULL if include_metadata else _WRAP_COLS_MINIMAL

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    _write_header(ws, headers)
    for row in quiz_to_rows(quiz, meta, include_metadata=include_metadata):
        ws.append(row)

    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if get_column_letter(cell.column) in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if metrics and isinstance(metrics, dict):
        _write_metrics_sheet(wb, metrics)

    wb.save(out_path)
    return out_path


def quiz_to_csv_rows(
    quiz: List[Dict[str, Any]],
    meta: Optional[ExcelMeta] = None,
    include_metadata: bool = True,
) -> Tuple[List[str], List[List[Any]]]:
    headers = HEADERS_FULL if include_metadata else HEADERS_MINIMAL
    return headers, quiz_to_rows(quiz, meta or ExcelMeta(), include_metadata=include_metadata)
